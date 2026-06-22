import os
import sys
import subprocess

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing it now...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_hosting_sheet():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styling helpers
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1E3A8A")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    small_italic_font = Font(name=font_family, size=9, italic=True, color="555555")
    
    # Fills
    primary_dark_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
    secondary_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")     # Royal Blue
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")          # Very light grey-blue
    
    # Borders
    thin_border_side = Side(style='thin', color='CBD5E1')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Alignments
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # -------------------------------------------------------------
    # SHEET 1: Hosting Cost Summary (Official Overview)
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Hosting Cost Summary")
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = "IIITDM ViBeS Lab - Website Hosting & Domain Cost Summary"
    title_cell.font = title_font
    title_cell.fill = primary_dark_fill
    title_cell.alignment = align_center
    ws1.row_dimensions[1].height = 40
    
    # Technical Architecture Context (Formal Info Block)
    ws1["A3"] = "System Architecture Context"
    ws1["A3"].font = section_font
    
    arch_specs = [
        ["Frontend Framework:", "TanStack Start (SSR Support Required)"],
        ["Backend API:", "Express.js / Node.js (REST API Server)"],
        ["Database:", "MongoDB Atlas (NoSQL Cloud Database)"],
    ]
    
    for idx, (spec_label, spec_val) in enumerate(arch_specs, start=4):
        ws1.cell(row=idx, column=1, value=spec_label).font = bold_font
        ws1.cell(row=idx, column=2, value=spec_val).font = regular_font
        ws1.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)
        ws1.row_dimensions[idx].height = 20
        
    # Table Title
    ws1["A8"] = "Hosting Budget Scenarios"
    ws1["A8"].font = section_font
    
    headers1 = [
        "Scenario", 
        "Frontend Host", 
        "Backend Host", 
        "Database Host", 
        "Domain Name & Registrar", 
        "Est. Monthly Cost (USD)", 
        "Est. Annual Cost (INR)"
    ]
    for col_idx, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=9, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = secondary_fill
        cell.alignment = align_header
        cell.border = thin_border
    ws1.row_dimensions[9].height = 28
    
    # Scenarios Data (Numeric values for monthly and formula for annual cost)
    # Annual cost formula: = (Monthly Cost * 12 + Domain Cost) * Exchange Rate (84)
    # Scenario 1 (Free tier): Domain = $5.75 (Porkbun vibeslab.in)
    # Scenario 2 (Managed): Domain = $9.75 (Porkbun vibeslab.com)
    # Scenario 3 (VPS): Domain = $5.75 (Porkbun vibeslab.in)
    # Scenario 4 (Scalable): Domain = $9.75 (Porkbun vibeslab.com)
    scenarios_data = [
        ["1. Free Tier (Development)", "Cloudflare Pages (Free)", "Render (Free - 15m sleep)", "MongoDB Atlas (Free M0)", "vibeslab.in (Porkbun)", 0.00, "=(F10*12 + 5.75)*84"],
        ["2. Managed PaaS (Recommended)", "Cloudflare Pages (Free)", "Railway (Developer)", "MongoDB Atlas (Free M0)", "vibeslab.com (Porkbun)", 5.00, "=(F11*12 + 9.75)*84"],
        ["3. Dedicated VPS (Full Control)", "Self-Hosted on VPS", "Self-Hosted on VPS", "MongoDB Atlas (Free M0)", "vibeslab.in (Porkbun)", 4.00, "=(F12*12 + 5.75)*84"],
        ["4. Fully Scalable Managed Stack", "Cloudflare Pages (Free)", "Render (Starter)", "MongoDB Atlas (Free M0)", "vibeslab.com (Porkbun)", 7.00, "=(F13*12 + 9.75)*84"],
    ]
    
    for r_idx, row_data in enumerate(scenarios_data, start=10):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx in [1, 2, 3, 4, 5]:
                cell.alignment = align_left
            elif c_idx == 6:
                cell.alignment = align_right
                cell.number_format = '"$"#,##0.00'
            elif c_idx == 7:
                cell.alignment = align_right
                cell.number_format = '"Rs. "#,##0'
            
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        ws1.row_dimensions[r_idx].height = 24
        


    # -------------------------------------------------------------
    # SHEET 2: Domain Registrars (Cost & Options)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Domain Registrars")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:I1")
    title2 = ws2["A1"]
    title2.value = "Domain Cost Comparison & Proposed Names (ViBeS Lab)"
    title2.font = title_font
    title2.fill = primary_dark_fill
    title2.alignment = align_center
    ws2.row_dimensions[1].height = 35
    
    # Section 1: Proposed Domain Names using 'vibes'
    ws2["A3"] = "Proposed Domain Name Options for ViBeS Lab"
    ws2["A3"].font = section_font
    
    headers_domains = ["Proposed Domain Name", "Type / Extension", "Primary Focus & Use Case"]
    for col_idx, h_name in enumerate(headers_domains, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h_name)
        cell.font = header_font
        cell.fill = secondary_fill
        cell.alignment = align_header
        cell.border = thin_border
    ws2.row_dimensions[4].height = 24
    
    proposed_domains = [
        ["vibeslab.in", "National (.in)", "Highly recommended for Indian research laboratories. Standard choice."],
        ["vibeslab.com", "Commercial (.com)", "Recommended for global reach and visibility. Industry standard."],
        ["vibes-lab.in", "National Alternative (.in)", "Backup choice if vibeslab.in is unavailable."],
        ["vibes-lab.com", "Commercial Alternative (.com)", "Backup choice if vibeslab.com is unavailable."],
        ["iiitdmvibes.in", "Institutional (.in)", "Directly highlights institutional affiliation with IIITDM."],
        ["vibeslab.org", "Organization (.org)", "Alternative extension highlighting the non-profit academic research nature."],
    ]
    
    for r_idx, row_data in enumerate(proposed_domains, start=5):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_left
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        ws2.row_dimensions[r_idx].height = 22
        
    # Section 2: Domain Registrar Comparison
    ws2["A12"] = "Registrar Cost Comparison"
    ws2["A12"].font = section_font
    
    headers2 = [
        "Registrar", 
        "Domain Extension", 
        "Example Domain", 
        "1st Year Cost (USD)", 
        "Renewal Cost (USD)", 
        "1st Year Cost (INR)", 
        "Renewal Cost (INR)", 
        "Privacy Protection", 
        "Best For"
    ]
    for col_idx, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=13, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = secondary_fill
        cell.alignment = align_header
        cell.border = thin_border
    ws2.row_dimensions[13].height = 28
    
    domain_data = [
        ["Porkbun", ".in (India)", "vibeslab.in", 5.75, 7.50, "=D14*84", "=E14*84", "Free Forever", "Cheapest .in domains, highly recommended"],
        ["Porkbun", ".com (Global)", "vibeslab.com", 9.75, 10.50, "=D15*84", "=E15*84", "Free Forever", "Excellent pricing, clean UI"],
        ["Namecheap", ".in (India)", "vibeslab.in", 6.98, 9.98, "=D16*84", "=E16*84", "Free Forever", "Reliable, very popular registrar"],
        ["Namecheap", ".com (Global)", "vibeslab.com", 8.98, 14.98, "=D17*84", "=E17*84", "Free Forever", "Industry standard, great promo deals"],
        ["Hostinger", ".in (India)", "vibeslab.in", 4.99, 8.99, "=D18*84", "=E18*84", "Free Forever", "Good localized payment options in India"],
        ["Hostinger", ".com (Global)", "vibeslab.com", 9.99, 13.99, "=D19*84", "=E19*84", "Free Forever", "Easy domain + web hosting package options"],
        ["GoDaddy", ".in (India)", "vibeslab.in", 4.99, 14.99, "=D20*84", "=E20*84", "Extra Charge ($9.99/yr)", "Expensive renewal rates, privacy is not free"],
        ["GoDaddy", ".com (Global)", "vibeslab.com", 11.99, 21.99, "=D21*84", "=E21*84", "Extra Charge ($9.99/yr)", "Expensive renewals, upsells everything"],
    ]
    
    for r_idx, row_data in enumerate(domain_data, start=14):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx in [1, 2, 3, 8, 9]:
                cell.alignment = align_left
            elif c_idx in [4, 5]:
                cell.alignment = align_right
                cell.number_format = '"$"#,##0.00'
            elif c_idx in [6, 7]:
                cell.alignment = align_right
                cell.number_format = '"Rs. "#,##0'
            
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        ws2.row_dimensions[r_idx].height = 24
        
    ws2["A23"] = "* Note: Prices exclude GST/taxes. USD to INR conversion rate is 1 USD = 84 INR."
    ws2["A23"].font = small_italic_font

    # -------------------------------------------------------------
    # SHEET 3: Frontend Hosting Options
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Frontend Hosting")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:G1")
    title3 = ws3["A1"]
    title3.value = "Frontend Static & Serverless Hosting Options"
    title3.font = title_font
    title3.fill = primary_dark_fill
    title3.alignment = align_center
    ws3.row_dimensions[1].height = 35
    
    headers3 = ["Provider", "Tier", "Monthly Price (USD)", "Monthly Price (INR)", "Bandwidth / Limits", "TanStack Start SSR Support", "Key Advantages"]
    for col_idx, header in enumerate(headers3, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = secondary_fill
        cell.alignment = align_header
        cell.border = thin_border
    ws3.row_dimensions[3].height = 28
    
    frontend_data = [
        ["Cloudflare Pages", "Free Tier", 0.00, "=C4*84", "Unlimited Bandwidth, 500 builds/month", "Yes (via Cloudflare Workers edge adapter)", "Extremely fast global edge CDN, unlimited bandwidth, free custom SSL."],
        ["Vercel", "Hobby", 0.00, "=C5*84", "100 GB Bandwidth, 6000 build minutes", "Yes (first class Vercel adapter)", "Easiest integration with GitHub, excellent developer tools and preview deployments."],
        ["Netlify", "Starter", 0.00, "=C6*84", "100 GB Bandwidth, 300 build minutes", "Yes (via Netlify Functions adapter)", "Great admin panel, easy DNS settings, form handling options built-in."],
        ["GitHub Pages", "Free", 0.00, "=C7*84", "100 GB Bandwidth", "No (Static files only, no SSR)", "Simple, but does not support TanStack Start's SSR features (only static builds)."],
    ]
    
    for r_idx, row_data in enumerate(frontend_data, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx in [1, 2, 5, 7]:
                cell.alignment = align_left
            elif c_idx == 3:
                cell.alignment = align_right
                cell.number_format = '"$"#,##0.00'
            elif c_idx == 4:
                cell.alignment = align_right
                cell.number_format = '"Rs. "#,##0'
            else:
                cell.alignment = align_center
            
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        ws3.row_dimensions[r_idx].height = 26

    # -------------------------------------------------------------
    # SHEET 4: Backend & Database Hosting Options
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Backend & DB Hosting")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.merge_cells("A1:H1")
    title4 = ws4["A1"]
    title4.value = "Backend Server (Express API) & Database (MongoDB) Hosting"
    title4.font = title_font
    title4.fill = primary_dark_fill
    title4.alignment = align_center
    ws4.row_dimensions[1].height = 35
    
    headers4 = ["Provider", "Type", "Plan / Specs", "Monthly Price (USD)", "Monthly Price (INR)", "Spin-down (Sleeps)?", "Pros", "Cons"]
    for col_idx, header in enumerate(headers4, start=1):
        cell = ws4.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = secondary_fill
        cell.alignment = align_header
        cell.border = thin_border
    ws4.row_dimensions[3].height = 28
    
    backend_data = [
        ["Render", "PaaS", "Free Web Service", 0.00, "=D4*84", "Yes (after 15m inactivity)", "Free, simple GitHub deployment, auto SSL", "50-second wake-up delay on first visit after sleep."],
        ["Render", "PaaS", "Starter (512MB RAM, 0.5 CPU)", 7.00, "=D5*84", "No", "Runs 24/7, zero setup management", "Can get expensive if adding custom disks or CPU."],
        ["Railway", "PaaS", "Developer Plan (Shared resource)", 5.00, "=D6*84", "No", "Extremely fast builds, runs 24/7, pay for usage", "Requires a valid card to activate the billing limit."],
        ["DigitalOcean", "VPS", "Basic VM (1GB RAM, 1 CPU, 25GB SSD)", 4.00, "=D7*84", "No", "Complete root server, cheap, can host multiple APIs", "Manual server setup (SSH, Nginx, PM2, Certbot required)."],
        ["Hetzner", "VPS", "CX22 Cloud (2GB RAM, 1 CPU, 40GB SSD)", 4.10, "=D8*84", "No", "Best performance-per-dollar, 2GB RAM", "Manual setup, server in Europe/US (higher ping in India)."],
        ["MongoDB Atlas", "Database", "Shared M0 Cluster (512MB storage)", 0.00, "=D9*84", "No", "Free forever, fully managed database backups", "Shared resources, limited to 100 read/write per sec."],
        ["MongoDB Atlas", "Database", "M2 Serverless (2GB storage)", 9.00, "=D10*84", "No", "Good for active scaling, daily automated backups", "Monthly cost can increase with high query traffic."],
    ]
    
    for r_idx, row_data in enumerate(backend_data, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx in [1, 2, 3, 7, 8]:
                cell.alignment = align_left
            elif c_idx == 4:
                cell.alignment = align_right
                cell.number_format = '"$"#,##0.00'
            elif c_idx == 5:
                cell.alignment = align_right
                cell.number_format = '"Rs. "#,##0'
            else:
                cell.alignment = align_center
            
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
        ws4.row_dimensions[r_idx].height = 26

    # -------------------------------------------------------------
    # Autofit Column Widths & Margins for all sheets
    # -------------------------------------------------------------
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Find max length of cells in column
            for cell in col:
                if cell.row == 1:  # Skip title row from calculation
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if val_str.startswith('='):
                        # Approximate length for formula results to avoid making columns narrow
                        max_len = max(max_len, 10)
                        continue
                    lines = val_str.split('\n')
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
            
            # Set width with a bit of padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to workspace root
    output_filename = "Hosting_and_Domain_Cost_Comparison.xlsx"
    workspace_root = r"c:\Users\vempa\OneDrive\Desktop\Lab Website"
    output_path = os.path.join(workspace_root, output_filename)
    
    wb.save(output_path)
    print(f"Excel sheet successfully created at: {output_path}")

if __name__ == "__main__":
    create_hosting_sheet()
